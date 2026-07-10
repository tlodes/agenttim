"""Regenerate Versuchsmatrix_AgentTim_v2.xlsx with current experiment config."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
wb = Workbook()
HF = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HB = PatternFill("solid", fgColor="4472C4")
SF = PatternFill("solid", fgColor="D9E2F3")
BF = Font(name="Arial", size=10)
TB = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
CHECK = "\u2713"
DASH = "\u2014"
def hdr(ws, row, cols):

    for i, h in enumerate(cols, 1):

        c = ws.cell(row=row, column=i, value=h)

        c.font = HF

        c.fill = HB

        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        c.border = TB
def body(ws, r1, r2, nc):

    for r in range(r1, r2 + 1):

        for c in range(1, nc + 1):

            cell = ws.cell(row=r, column=c)

            cell.font = BF

            cell.border = TB

            if c > 1:

                cell.alignment = Alignment(horizontal="center")
def sec(ws, row, nc, label):

    ws.cell(row=row, column=1, value=label)

    for c in range(1, nc + 1):

        cell = ws.cell(row=row, column=c)

        cell.fill = SF

        cell.font = Font(name="Arial", bold=True, size=10)

        cell.border = TB
ws = wb.active
ws.title = "Evaluation Matrix"
ws["A1"] = "AgentTim Versuchsmatrix"
ws["A1"].font = Font(name="Arial", bold=True, size=14)
ws["A2"] = "Masterarbeit: Multi-Agent Pattern Comparison | Model: GPT-5.4-mini"
ws["A2"].font = Font(name="Arial", size=10, color="666666")
cols = ["Benchmark / Mode", "Single Agent", "Orchestrator (Fine)", "Orchestrator (Coarse)", "Router", "Swarm"]
hdr(ws, 4, cols)
nc = len(cols)
data = [

    (True, "MCPAgentBench (Single-Turn, 178 cases)"),

    (False, "  STurn-1T daytasks (30)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  STurn-1T protasks (30)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  STurn-MT daytasks parallel (20)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  STurn-MT daytasks sequential (20)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  STurn-MT daytasks 3tools (20)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  STurn-MT protasks parallel (20)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  STurn-MT protasks sequential (20)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  STurn-MT protasks 3tools (18)", CHECK, CHECK, CHECK, CHECK, CHECK),

    (True, "BFCL (Multi-Turn, 200 cases)"),

    (False, "  MTurn-MT base (100)", CHECK, CHECK, DASH, CHECK, CHECK),

    (False, "  MTurn-MT long_context (100)", CHECK, CHECK, DASH, CHECK, CHECK),

    (True, "Tool Count Variations (per benchmark)"),

    (False, "  --num-tools 10", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  --num-tools 20", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  --num-tools 40", CHECK, CHECK, CHECK, CHECK, CHECK),

    (False, "  --num-tools 80", CHECK, CHECK, CHECK, CHECK, CHECK),
]
r = 5
for row in data:

    is_sec = row[0]

    if is_sec:

        sec(ws, r, nc, row[1])

    else:

        ws.cell(row=r, column=1, value=row[1])

        for c, v in enumerate(row[2:], 2):

            ws.cell(row=r, column=c, value=v)

    r += 1
body(ws, 5, r - 1, nc)
ws.column_dimensions["A"].width = 42
for c in range(2, nc + 1):

    ws.column_dimensions[get_column_letter(c)].width = 22
ws2 = wb.create_sheet("Metrics")
ws2["A1"] = "Evaluation Metrics"
ws2["A1"].font = Font(name="Arial", bold=True, size=14)
cols2 = ["Metric", "Type", "Benchmark", "Description"]
hdr(ws2, 3, cols2)
metrics = [

    ("Tool Correctness", "Core", "Both", "Subset match: expected tools present in actual"),

    ("Argument Correctness [Reference]", "Core", "Both", "LLM-as-judge (GPT-4.1-mini) comparing expected vs actual args"),

    ("Final State (Derived)", "Core", "MCPAgentBench", "Binary: TC=1.0 AND AC=1.0 (ignores execution order)"),

    ("State Match", "Core", "BFCL", "Final Python class state == ground truth state"),

    ("Tool Correctness (per-turn)", "Core", "BFCL", "Per-turn avg + cross-turn redundancy detection"),

    ("Execution Efficiency", "Informational", "MCPAgentBench (parallel)", "Timestamp overlap analysis for parallel tools"),

    ("Coordination Token Overhead", "Post-hoc", "Both (multi-agent)", "(multi_tokens - single_tokens) / single_tokens"),

    ("Token Usage / Cost", "Tracked", "Both", "Prompt + completion tokens x model pricing"),

    ("Latency", "Tracked", "Both", "Wall-clock time per test case"),
]
for i, row in enumerate(metrics, 4):

    for c, v in enumerate(row, 1):

        ws2.cell(row=i, column=c, value=v)
body(ws2, 4, 4 + len(metrics) - 1, len(cols2))
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 58
ws3 = wb.create_sheet("Agent Architectures")
ws3["A1"] = "Agent Architecture Comparison"
ws3["A1"].font = Font(name="Arial", bold=True, size=14)
cols3 = ["Characteristic", "Single Agent", "Orchestrator", "Router", "Swarm"]
hdr(ws3, 3, cols3)
arch = [

    ("Interaction Type", "Agent <-> Tools", "Supervisor -> Subagents -> Tools", "Classifier -> Subagents -> Synth.", "Analyst -> Planner -> Critic -> Exec."),

    ("LLM Calls / Task", "O(k)", "O(k + n*k')", "O(1 + k' + 1)", "O(3 + k')"),

    ("Communication Overhead", "0", "n delegations + n results", "1 classification + n results", "3 handoffs"),

    ("Parallelization", "No", "No (sequential)", "Yes (Send())", "No (pipeline)"),

    ("Tool Visibility", "All tools", "Supervisor: delegation\nSubagents: domain", "Same as Orchestrator", "Planner: text\nExecutor: all tools"),

    ("Context per Agent", "Full history", "Supervisor: full\nSubagents: task string", "Classify: query\nSubagents: sub-query", "Each role: prior outputs"),

    ("Routing", "ReAct loop", "ReAct (iterative)", "Structured output (1-shot)", "Fixed pipeline"),

    ("Granularity", "N/A", "Fine (16/8) + Coarse (3)", "Fine only", "N/A"),

    ("Recursion Limit", "20", "20 (supervisor) + 20 (subagent)", "20", "20 (executor)"),

    ("Base Class", "ReactEvalAgent", "ReactEvalAgent", "StatefulEvalAgent", "StatefulEvalAgent"),
]
for i, row in enumerate(arch, 4):

    for c, v in enumerate(row, 1):

        ws3.cell(row=i, column=c, value=v)
body(ws3, 4, 4 + len(arch) - 1, len(cols3))
ws3.column_dimensions["A"].width = 24
for c in range(2, len(cols3) + 1):

    ws3.column_dimensions[get_column_letter(c)].width = 34
ws4 = wb.create_sheet("Configuration")
ws4["A1"] = "Experiment Configuration"
ws4["A1"].font = Font(name="Arial", bold=True, size=14)
cols4 = ["Parameter", "Value", "Rationale"]
hdr(ws4, 3, cols4)
conf = [

    ("Agent Model", "GPT-5.4-mini (gpt-5.4-mini)", "Frontier mini model, cost-effective"),

    ("Judge Model", "GPT-4.1-mini (june-gpt-4-1-mini-datazone)", "Cheaper, sufficient for argument comparison"),

    ("Temperature", "0.1", "Near-deterministic, reproducible"),

    ("Recursion Limit", "20 (uniform, all agents)", "Matches BFCL MAXIMUM_STEP_LIMIT=20"),

    ("Turn Timeout", "120s", "Hard timeout per invocation"),

    ("Parallel Tool Calls", "Enabled", "Multiple tool calls per LLM response"),

    ("Tool Binding", "OpenAI function-calling", "Not prompt-based"),

    ("Tool Counts", "10, 20, 40, 80", "Variable distractor count per test case"),

    ("System Prompt", "BASE_PROMPT per benchmark", "From original papers, adapted"),

    ("Pricing Input/1M", "$0.75", "GPT-5.4-mini short context"),

    ("Pricing Output/1M", "$4.50", "GPT-5.4-mini short context"),
]
for i, row in enumerate(conf, 4):

    for c, v in enumerate(row, 1):

        ws4.cell(row=i, column=c, value=v)
body(ws4, 4, 4 + len(conf) - 1, len(cols4))
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 38
ws4.column_dimensions["C"].width = 50
ws5 = wb.create_sheet("Benchmarks")
ws5["A1"] = "Benchmark Comparison"
ws5["A1"].font = Font(name="Arial", bold=True, size=14)
cols5 = ["", "MCPAgentBench", "BFCL"]
hdr(ws5, 3, cols5)
bench = [

    ("Source", "Liu et al., 2025", "Berkeley FC Leaderboard"),

    ("Turn Type", "Single-turn", "Multi-turn (2-5 turns)"),

    ("Total Tools", "141", "128"),

    ("Domains", "16", "8"),

    ("Tool Behavior", "Deterministic mock data", "Stateful Python classes"),

    ("State Management", "None", "MCP reset + GT replay"),

    ("Original Prompt (FC mode)", "Yes", "No"),

    ("Our System Prompt", "BASE_PROMPT (adapted)", "BASE_PROMPT (adapted) + cd hint"),

    ("Primary Metric", "Final State (Derived)", "State Match"),

    ("Test Cases", "178", "200"),

    ("Modes", "8 (day/pro x 1t/par/seq/3t)", "2 (base, long_context)"),
]
for i, row in enumerate(bench, 4):

    for c, v in enumerate(row, 1):

        ws5.cell(row=i, column=c, value=v)
body(ws5, 4, 4 + len(bench) - 1, len(cols5))
ws5.column_dimensions["A"].width = 28
ws5.column_dimensions["B"].width = 35
ws5.column_dimensions["C"].width = 35
out = Path(__file__).resolve().parent.parent / "Versuchsmatrix_AgentTim_v2.xlsx"
wb.save(out)
print(f"Saved to {out}")

